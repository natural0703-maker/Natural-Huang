from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.worksheet import _writer as worksheet_writer

from src.models import (
    AnomalyRecord,
    FailureRecord,
    ReplacementRecord,
    ReviewCandidateRecord,
    SummaryRecord,
)


def write_report(
    report_path: Path,
    summaries: list[SummaryRecord],
    replacement_records: list[ReplacementRecord],
    review_candidates: list[ReviewCandidateRecord],
    anomalies: list[AnomalyRecord],
    failures: list[FailureRecord],
) -> Path:
    ordered_review_candidates = sorted(
        review_candidates,
        key=lambda item: (item.file_name, item.paragraph_index, item.candidate_id),
    )
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(
        [
            "file_name",
            "paragraph_count",
            "total_replacements",
            "total_review_candidates",
            "review_grammar_count",
            "review_wording_count",
            "review_regional_usage_count",
            "total_anomalies",
            "status",
            "elapsed_time_sec",
            "output_file",
        ]
    )
    for summary in summaries:
        summary_sheet.append(
            [
                summary.file_name,
                summary.paragraph_count,
                summary.total_replacements,
                summary.total_review_candidates,
                summary.review_grammar_count,
                summary.review_wording_count,
                summary.review_regional_usage_count,
                summary.total_anomalies,
                summary.status,
                summary.elapsed_time_sec,
                summary.output_file,
            ]
        )

    replacements_sheet = workbook.create_sheet("replacements")
    replacements_sheet.append(
        [
            "file_name",
            "paragraph_index",
            "original_snippet",
            "replaced_term",
            "target_term",
            "replacement_count",
        ]
    )
    for record in replacement_records:
        replacements_sheet.append(
            [
                record.file_name,
                record.paragraph_index,
                record.original_snippet,
                record.replaced_term,
                record.target_term,
                record.replacement_count,
            ]
        )

    review_sheet = workbook.create_sheet("review_candidates")
    review_sheet.append(
        [
            "candidate_id",
            "file_name",
            "chapter_guess",
            "paragraph_index",
            "position_hint",
            "hit_term",
            "risk_category",
            "original_snippet",
            "processed_snippet",
            "context_snippet",
            "suggested_candidates",
            "confidence",
            "status",
            "note",
            "resolved_text",
        ]
    )
    for candidate in ordered_review_candidates:
        review_sheet.append(
            [
                candidate.candidate_id,
                candidate.file_name,
                candidate.chapter_guess,
                candidate.paragraph_index,
                candidate.position_hint,
                candidate.hit_term,
                candidate.risk_category,
                candidate.original_snippet,
                candidate.processed_snippet,
                candidate.context_snippet,
                candidate.suggested_candidates,
                candidate.confidence,
                candidate.status,
                candidate.note,
                candidate.resolved_text,
            ]
        )

    anomalies_sheet = workbook.create_sheet("anomalies")
    anomalies_sheet.append(
        [
            "file_name",
            "paragraph_index",
            "anomaly_char",
            "original_snippet",
            "converted_snippet",
        ]
    )
    for anomaly in anomalies:
        anomalies_sheet.append(
            [
                anomaly.file_name,
                anomaly.paragraph_index,
                anomaly.anomaly_char,
                anomaly.original_snippet,
                anomaly.converted_snippet,
            ]
        )

    failures_sheet = workbook.create_sheet("failures")
    failures_sheet.append(["file_name", "error_type", "error_message"])
    for failure in failures:
        failures_sheet.append([failure.file_name, failure.error_type, failure.error_message])

    save_workbook_safely(workbook, report_path)
    review_summary_path = report_path.with_name("review_summary.xlsx")
    _write_review_summary(review_summary_path, ordered_review_candidates)
    return review_summary_path


def save_workbook_safely(workbook: Workbook, report_path: Path) -> None:
    try:
        workbook.save(str(report_path))
        return
    except FileNotFoundError:
        # Fallback for environments where system TEMP cannot be used reliably.
        pass

    temp_root = report_path.parent / ".xlsx_tmp"
    temp_root.mkdir(parents=True, exist_ok=True)

    original_create_tmp = worksheet_writer.create_temporary_file
    original_cleanup = worksheet_writer.WorksheetWriter.cleanup

    def _create_local_tmp(suffix: str = "") -> str:
        temp_file = temp_root / f"sheet_{uuid4().hex}{suffix}"
        temp_file.touch(exist_ok=True)
        return str(temp_file)

    def _cleanup_noop(self) -> None:  # type: ignore[no-untyped-def]
        return None

    worksheet_writer.create_temporary_file = _create_local_tmp
    worksheet_writer.WorksheetWriter.cleanup = _cleanup_noop
    try:
        workbook.save(str(report_path))
    finally:
        worksheet_writer.create_temporary_file = original_create_tmp
        worksheet_writer.WorksheetWriter.cleanup = original_cleanup


def _write_review_summary(
    review_summary_path: Path,
    review_candidates: list[ReviewCandidateRecord],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review_summary"
    sheet.append(
        [
            "candidate_id",
            "file_name",
            "chapter_guess",
            "paragraph_index",
            "hit_term",
            "risk_category",
            "context_snippet",
            "suggested_candidates",
            "confidence",
            "status",
            "note",
            "resolved_text",
        ]
    )
    for item in review_candidates:
        sheet.append(
            [
                item.candidate_id,
                item.file_name,
                item.chapter_guess,
                item.paragraph_index,
                item.hit_term,
                item.risk_category,
                item.context_snippet,
                item.suggested_candidates,
                item.confidence,
                item.status,
                item.note,
                item.resolved_text,
            ]
        )
    save_workbook_safely(workbook, review_summary_path)
