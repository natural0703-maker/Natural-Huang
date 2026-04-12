from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.config_loader import DocumentFormatConfig
from src.docx_reader import read_paragraphs
from src.docx_writer import ParagraphOutput, build_reviewed_output_docx_path, write_paragraphs_to_docx
from src.heading_detector import is_direct_chapter_heading
from src.models import (
    ApplyDecisionResultRecord,
    ApplyFailureRecord,
    ApplySummaryRecord,
    ReviewDecisionRecord,
)
from src.report_writer import save_workbook_safely


ALLOWED_ACCEPT_STATUS = "accepted"
STATUS_VALUES = {"pending", "accepted", "rejected", "skip"}


@dataclass(frozen=True)
class ApplyRunResult:
    summaries: list[ApplySummaryRecord]
    decision_results: list[ApplyDecisionResultRecord]
    failures: list[ApplyFailureRecord]
    apply_summary_path: Path
    output_files: list[Path]
    total_candidates: int
    applied_count: int
    skipped_count: int
    not_found_count: int
    conflict_count: int
    failed_count: int
    reason_counts: dict[str, int]


def load_review_decisions(summary_path: Path) -> list[ReviewDecisionRecord]:
    suffix = summary_path.suffix.lower()
    if suffix == ".xlsx":
        return _load_review_decisions_xlsx(summary_path)
    if suffix == ".csv":
        return _load_review_decisions_csv(summary_path)
    raise ValueError(f"不支援的 review_summary 格式: {summary_path}")


def apply_review_decisions(
    source_files: list[Path],
    output_dir: Path,
    summary_path: Path,
    document_format: DocumentFormatConfig | None = None,
) -> ApplyRunResult:
    decisions = load_review_decisions(summary_path)
    decisions_by_file: dict[str, list[ReviewDecisionRecord]] = defaultdict(list)
    for item in decisions:
        decisions_by_file[_normalize_file_key(item.file_name)].append(item)

    summaries: list[ApplySummaryRecord] = []
    decision_results: list[ApplyDecisionResultRecord] = []
    failures: list[ApplyFailureRecord] = []
    output_files: list[Path] = []
    reason_counts: dict[str, int] = defaultdict(int)

    total_candidates = 0
    applied_count = 0
    skipped_count = 0
    not_found_count = 0
    conflict_count = 0
    failed_count = 0

    matched_file_keys: set[str] = set()
    for source_file in source_files:
        matched_key = _normalize_file_key(source_file.name)
        matched_file_keys.add(matched_key)
        file_decisions = decisions_by_file.get(_normalize_file_key(source_file.name), [])
        file_result = _apply_for_single_file(
            source_file=source_file,
            output_dir=output_dir,
            decisions=file_decisions,
            document_format=document_format,
        )
        summaries.append(file_result["summary"])
        decision_results.extend(file_result["decisions"])
        failures.extend(file_result["failures"])
        if file_result["output_file"] is not None:
            output_files.append(file_result["output_file"])

        total_candidates += file_result["summary"].total_candidates
        applied_count += file_result["summary"].applied_count
        skipped_count += file_result["summary"].skipped_count
        not_found_count += file_result["summary"].not_found_count
        conflict_count += file_result["summary"].conflict_count
        failed_count += file_result["summary"].failed_count
        for key, value in file_result["reason_counts"].items():
            reason_counts[key] = reason_counts.get(key, 0) + int(value)

    for file_key, unmatched_decisions in decisions_by_file.items():
        if file_key in matched_file_keys:
            continue
        for item in unmatched_decisions:
            failed_count += 1
            reason_counts["target file mismatch"] = reason_counts.get("target file mismatch", 0) + 1
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=item.file_name,
                    candidate_id=item.candidate_id,
                    paragraph_index=item.paragraph_index,
                    status=item.status,
                    applied=False,
                    reason="target file mismatch",
                    hit_term=item.hit_term,
                    resolved_text=item.resolved_text,
                )
            )

    apply_summary_path = output_dir / "apply_summary.xlsx"
    _write_apply_summary(
        apply_summary_path=apply_summary_path,
        summaries=summaries,
        decision_results=decision_results,
        failures=failures,
        reason_counts=dict(reason_counts),
    )

    return ApplyRunResult(
        summaries=summaries,
        decision_results=decision_results,
        failures=failures,
        apply_summary_path=apply_summary_path,
        output_files=output_files,
        total_candidates=total_candidates,
        applied_count=applied_count,
        skipped_count=skipped_count,
        not_found_count=not_found_count,
        conflict_count=conflict_count,
        failed_count=failed_count,
        reason_counts=dict(reason_counts),
    )


def _normalize_file_key(file_name: str) -> str:
    lowered = file_name.strip().lower()
    if lowered.endswith(".docx"):
        lowered = lowered[:-5]
    for suffix in ("_reviewed", "_tw"):
        if lowered.endswith(suffix):
            lowered = lowered[: -len(suffix)]
            break
    return lowered


def _apply_for_single_file(
    source_file: Path,
    output_dir: Path,
    decisions: list[ReviewDecisionRecord],
    document_format: DocumentFormatConfig | None,
) -> dict[str, object]:
    paragraphs = read_paragraphs(source_file)
    updated_paragraphs = list(paragraphs)

    candidate_total = len(decisions)
    applied = 0
    skipped = 0
    not_found = 0
    conflict = 0
    failed = 0
    reason_counts: dict[str, int] = defaultdict(int)

    decision_results: list[ApplyDecisionResultRecord] = []
    failures: list[ApplyFailureRecord] = []
    seen_candidate_ids: set[str] = set()
    original_paragraphs = list(paragraphs)

    for decision in decisions:
        if decision.candidate_id in seen_candidate_ids:
            conflict += 1
            reason_counts["multiple matches"] = reason_counts.get("multiple matches", 0) + 1
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="multiple matches",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
            continue
        seen_candidate_ids.add(decision.candidate_id)

        normalized_status = decision.status.strip().lower()
        resolved_text = decision.resolved_text.strip()
        if normalized_status not in STATUS_VALUES:
            skipped += 1
            reason_counts["no accepted resolved_text"] = (
                reason_counts.get("no accepted resolved_text", 0) + 1
            )
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="no accepted resolved_text",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
            continue

        if normalized_status != ALLOWED_ACCEPT_STATUS or not resolved_text:
            skipped += 1
            reason_counts["no accepted resolved_text"] = (
                reason_counts.get("no accepted resolved_text", 0) + 1
            )
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="no accepted resolved_text",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
            continue

        if decision.paragraph_index < 1 or decision.paragraph_index > len(updated_paragraphs):
            not_found += 1
            reason_counts["paragraph not found"] = reason_counts.get("paragraph not found", 0) + 1
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="paragraph not found",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
            continue

        index = decision.paragraph_index - 1
        paragraph_text = updated_paragraphs[index]
        original_paragraph_text = original_paragraphs[index]
        apply_result = _apply_to_paragraph(
            paragraph_text=paragraph_text,
            original_paragraph_text=original_paragraph_text,
            hit_term=decision.hit_term,
            context_snippet=decision.context_snippet,
            resolved_text=resolved_text,
        )
        if apply_result == "applied":
            updated_paragraphs[index] = paragraph_text.replace(decision.hit_term, resolved_text, 1)
            applied += 1
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=True,
                    reason="applied",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
        elif apply_result == "not_found":
            not_found += 1
            reason_counts["context mismatch"] = reason_counts.get("context mismatch", 0) + 1
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="context mismatch",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
        elif apply_result == "conflict":
            conflict += 1
            reason_counts["multiple matches"] = reason_counts.get("multiple matches", 0) + 1
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="multiple matches",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )
        else:
            failed += 1
            reason_counts["other"] = reason_counts.get("other", 0) + 1
            failure = ApplyFailureRecord(
                file_name=source_file.name,
                candidate_id=decision.candidate_id,
                error_type="ApplyError",
                error_message=apply_result,
            )
            failures.append(failure)
            decision_results.append(
                ApplyDecisionResultRecord(
                    file_name=source_file.name,
                    candidate_id=decision.candidate_id,
                    paragraph_index=decision.paragraph_index,
                    status=decision.status,
                    applied=False,
                    reason="failed",
                    hit_term=decision.hit_term,
                    resolved_text=decision.resolved_text,
                )
            )

    output_file: Path | None = None
    if candidate_total > 0:
        output_file = build_reviewed_output_docx_path(source_file, output_dir)
        outputs = [
            ParagraphOutput(text=item, is_heading=is_direct_chapter_heading(item))
            for item in updated_paragraphs
        ]
        write_paragraphs_to_docx(outputs, output_file, document_format=document_format)

    summary = ApplySummaryRecord(
        file_name=source_file.name,
        source_file=str(source_file),
        output_file=str(output_file) if output_file else "",
        total_candidates=candidate_total,
        applied_count=applied,
        skipped_count=skipped,
        not_found_count=not_found,
        conflict_count=conflict,
        failed_count=failed,
    )
    return {
        "summary": summary,
        "decisions": decision_results,
        "failures": failures,
        "output_file": output_file,
        "reason_counts": dict(reason_counts),
    }


def _apply_to_paragraph(
    paragraph_text: str,
    original_paragraph_text: str,
    hit_term: str,
    context_snippet: str,
    resolved_text: str,
) -> str:
    _ = resolved_text
    hit_count = paragraph_text.count(hit_term)
    if hit_count == 0:
        return "not_found"
    if hit_count > 1:
        return "conflict"
    if context_snippet and context_snippet not in original_paragraph_text:
        return "not_found"
    return "applied"


def _load_review_decisions_xlsx(path: Path) -> list[ReviewDecisionRecord]:
    workbook = load_workbook(path)
    sheet_name = "review_summary" if "review_summary" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}

    required = ["candidate_id", "file_name", "paragraph_index", "hit_term", "context_snippet", "status", "resolved_text"]
    for item in required:
        if item not in header_index:
            raise ValueError(f"review_summary 缺少必要欄位: {item}")

    decisions: list[ReviewDecisionRecord] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        file_name = _cell_value(row, header_index, "file_name")
        if not file_name:
            continue
        decisions.append(
            ReviewDecisionRecord(
                candidate_id=_cell_value(row, header_index, "candidate_id"),
                file_name=file_name,
                paragraph_index=_to_int(_cell_value(row, header_index, "paragraph_index")),
                hit_term=_cell_value(row, header_index, "hit_term"),
                context_snippet=_cell_value(row, header_index, "context_snippet"),
                status=_cell_value(row, header_index, "status", "pending"),
                resolved_text=_cell_value(row, header_index, "resolved_text"),
                note=_cell_value(row, header_index, "note"),
                chapter_guess=_cell_value(row, header_index, "chapter_guess"),
                risk_category=_cell_value(row, header_index, "risk_category"),
                suggested_candidates=_cell_value(row, header_index, "suggested_candidates"),
                confidence=_to_float(_cell_value(row, header_index, "confidence")),
            )
        )
    return decisions


def _load_review_decisions_csv(path: Path) -> list[ReviewDecisionRecord]:
    decisions: list[ReviewDecisionRecord] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        required = ["candidate_id", "file_name", "paragraph_index", "hit_term", "context_snippet", "status", "resolved_text"]
        for item in required:
            if item not in (reader.fieldnames or []):
                raise ValueError(f"review_summary 缺少必要欄位: {item}")

        for row in reader:
            file_name = (row.get("file_name") or "").strip()
            if not file_name:
                continue
            decisions.append(
                ReviewDecisionRecord(
                    candidate_id=(row.get("candidate_id") or "").strip(),
                    file_name=file_name,
                    paragraph_index=_to_int((row.get("paragraph_index") or "").strip()),
                    hit_term=(row.get("hit_term") or "").strip(),
                    context_snippet=(row.get("context_snippet") or "").strip(),
                    status=(row.get("status") or "pending").strip(),
                    resolved_text=(row.get("resolved_text") or "").strip(),
                    note=(row.get("note") or "").strip(),
                    chapter_guess=(row.get("chapter_guess") or "").strip(),
                    risk_category=(row.get("risk_category") or "").strip(),
                    suggested_candidates=(row.get("suggested_candidates") or "").strip(),
                    confidence=_to_float((row.get("confidence") or "").strip()),
                )
            )
    return decisions


def _cell_value(
    row: tuple[object, ...],
    header_index: dict[str, int],
    key: str,
    default: str = "",
) -> str:
    idx = header_index.get(key)
    if idx is None or idx >= len(row):
        return default
    value = row[idx]
    if value is None:
        return default
    return str(value).strip()


def _to_int(value: str) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_float(value: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _write_apply_summary(
    apply_summary_path: Path,
    summaries: list[ApplySummaryRecord],
    decision_results: list[ApplyDecisionResultRecord],
    failures: list[ApplyFailureRecord],
    reason_counts: dict[str, int],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "apply_summary"
    summary_sheet.append(
        [
            "file_name",
            "source_file",
            "output_file",
            "total_candidates",
            "applied_count",
            "skipped_count",
            "not_found_count",
            "conflict_count",
            "failed_count",
        ]
    )
    for item in summaries:
        summary_sheet.append(
            [
                item.file_name,
                item.source_file,
                item.output_file,
                item.total_candidates,
                item.applied_count,
                item.skipped_count,
                item.not_found_count,
                item.conflict_count,
                item.failed_count,
            ]
        )

    decision_sheet = workbook.create_sheet("apply_details")
    decision_sheet.append(
        [
            "file_name",
            "candidate_id",
            "paragraph_index",
            "status",
            "applied",
            "reason",
            "reason_category",
            "hit_term",
            "resolved_text",
        ]
    )
    for item in decision_results:
        decision_sheet.append(
            [
                item.file_name,
                item.candidate_id,
                item.paragraph_index,
                item.status,
                item.applied,
                item.reason,
                item.reason,
                item.hit_term,
                item.resolved_text,
            ]
        )

    failure_sheet = workbook.create_sheet("apply_failures")
    failure_sheet.append(["file_name", "candidate_id", "error_type", "error_message"])
    for item in failures:
        failure_sheet.append([item.file_name, item.candidate_id, item.error_type, item.error_message])

    reason_sheet = workbook.create_sheet("apply_reason_stats")
    reason_sheet.append(["reason_category", "count"])
    for key in sorted(reason_counts):
        reason_sheet.append([key, int(reason_counts[key])])

    save_workbook_safely(workbook, apply_summary_path)
