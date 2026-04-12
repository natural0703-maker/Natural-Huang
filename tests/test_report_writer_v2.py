from pathlib import Path
from uuid import uuid4
from unittest.mock import patch

from openpyxl import load_workbook

from src.models import SummaryRecord
from src.report_writer import write_report
from tests.test_paths import make_test_dir


def test_report_contains_v2_sheets() -> None:
    tmp_path = make_test_dir("report_sheets")
    report_path = tmp_path / "report.xlsx"
    summaries = [
        SummaryRecord(
            file_name="a.docx",
            status="success",
            paragraph_count=10,
            total_replacements=3,
            total_review_candidates=2,
            total_anomalies=1,
            elapsed_time_sec=0.5,
            output_file="C:/out/a_TW.docx",
        )
    ]

    def fake_create_temporary_file(suffix: str = "") -> str:
        temp_file = tmp_path / f"openpyxl_tmp_{uuid4().hex}{suffix}"
        temp_file.touch()
        return str(temp_file)

    with patch("openpyxl.worksheet._writer.create_temporary_file", side_effect=fake_create_temporary_file), patch(
        "openpyxl.worksheet._writer.WorksheetWriter.cleanup", return_value=None
    ):
        write_report(
            report_path=report_path,
            summaries=summaries,
            replacement_records=[],
            review_candidates=[],
            anomalies=[],
            failures=[],
        )

    workbook = load_workbook(report_path)
    assert "summary" in workbook.sheetnames
    assert "replacements" in workbook.sheetnames
    assert "review_candidates" in workbook.sheetnames
    assert "anomalies" in workbook.sheetnames
    assert "failures" in workbook.sheetnames
