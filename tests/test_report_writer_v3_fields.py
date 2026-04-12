from openpyxl import load_workbook

from src.models import ReviewCandidateRecord, SummaryRecord
from src.report_writer import write_report
from tests.test_paths import make_test_dir


def test_review_candidates_sheet_has_v33_columns_and_review_summary() -> None:
    tmp = make_test_dir("report_review_columns_v33")
    report_path = tmp / "report.xlsx"

    summaries = [
        SummaryRecord(
            file_name="a.docx",
            status="success",
            paragraph_count=1,
            total_replacements=0,
            total_review_candidates=1,
            total_anomalies=0,
            elapsed_time_sec=0.01,
            output_file="C:/out/a_TW.docx",
            review_grammar_count=0,
            review_wording_count=1,
            review_regional_usage_count=0,
        )
    ]
    review_candidates = [
        ReviewCandidateRecord(
            file_name="a.docx",
            paragraph_index=1,
            hit_term="資訊",
            risk_category="wording",
            original_snippet="原文資訊片段",
            processed_snippet="處理後資訊片段",
            context_snippet="上下文資訊片段",
            suggested_candidates="建議改為資訊/訊息",
            confidence=0.55,
            candidate_id="a.docx-P00001-C001",
            chapter_guess="第一章 測試",
            position_hint="段落 1（第 1 處）",
            status="pending",
            note="",
            resolved_text="",
        )
    ]

    review_summary_path = write_report(
        report_path=report_path,
        summaries=summaries,
        replacement_records=[],
        review_candidates=review_candidates,
        anomalies=[],
        failures=[],
    )

    wb = load_workbook(report_path)
    ws = wb["review_candidates"]
    headers = [cell.value for cell in ws[1]]

    assert "candidate_id" in headers
    assert "chapter_guess" in headers
    assert "position_hint" in headers
    assert "risk_category" in headers
    assert "original_snippet" in headers
    assert "processed_snippet" in headers
    assert "status" in headers
    assert "note" in headers
    assert "resolved_text" in headers

    assert review_summary_path.exists()
    review_wb = load_workbook(review_summary_path)
    assert "review_summary" in review_wb.sheetnames
